#!/usr/bin/env python3
"""HUD measurement software socket client.

Switch one or more configuration files and execute measurement commands.
"""

from __future__ import annotations

import argparse
import ast
import re
import shutil
import socket
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from brightness_io import (
    BRIGHTNESS_HEIGHT,
    BRIGHTNESS_WIDTH,
    FLOAT32_BYTES,
    convert_brightness_bin_to_txt,
    create_corrected_brightness_txt,
)


TERMINATOR = b"%"


class HudProtocolError(RuntimeError):
    """Raised when the HUD software returns an unexpected response."""


@dataclass(frozen=True)
class TestResult:
    config: str
    command: str
    response: str
    image_path: str | None = None
    display_name: str | None = None


@dataclass(frozen=True)
class PlanOptions:
    sparkle_source_dir: str | None = None
    sparkle_save_dir: str | None = None


def parse_t24_rows(response: str) -> list[list[float]]:
    """Extract only column 2 from each comma-separated seven-column t24 row."""
    match = re.match(r"^t24_Result(?:\(\d+\))?:(.*?)%?$", response.strip(), re.DOTALL)
    if not match:
        raise HudProtocolError(f"Invalid t24 response format: {response!r}")

    groups = [group.strip() for group in match.group(1).split(",") if group.strip()]
    if not groups:
        raise HudProtocolError("t24 returned no data rows")

    rows: list[list[float]] = []
    for index, group in enumerate(groups, start=1):
        parts = group.split()
        if len(parts) != 7:
            raise HudProtocolError(
                f"t24 row {index} has {len(parts)} columns; expected 7: {group!r}"
            )
        try:
            rows.append([float(parts[1])])
        except ValueError as exc:
            raise HudProtocolError(
                f"t24 row {index} has a non-numeric second-column value: {parts[1]!r}"
            ) from exc
    return rows


def parse_t6_rows(response: str) -> list[list[float]]:
    """Extract only column 2 from each comma/newline-separated t6 data row."""
    match = re.match(r"^t6_Result(?:\(\d+\))?:(.*?)%?$", response.strip(), re.DOTALL)
    if not match:
        raise HudProtocolError(f"Invalid t6 response format: {response!r}")

    groups = [
        group.strip()
        for group in re.split(r",|\r?\n", match.group(1))
        if group.strip()
    ]
    if not groups:
        raise HudProtocolError("t6 returned no data rows")

    rows: list[list[float]] = []
    for index, group in enumerate(groups, start=1):
        parts = group.split()
        if len(parts) < 2:
            raise HudProtocolError(f"t6 row {index} has no second column: {group!r}")
        try:
            rows.append([float(parts[1])])
        except ValueError as exc:
            raise HudProtocolError(
                f"t6 row {index} has a non-numeric second-column value: {parts[1]!r}"
            ) from exc
    return rows


def export_results_excel(results: list[TestResult], output_path: str) -> Path:
    """Export t24/t6 column-2 values into separate measurement sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HudProtocolError(
            "Excel export requires openpyxl; install it with: python3 -m pip install openpyxl"
        ) from exc

    supported = {"t24": parse_t24_rows, "t6": parse_t6_rows}
    grouped: dict[str, list[tuple[str, list[list[float]]]]] = {}
    indexes: dict[str, dict[str, list[list[float]]]] = {}
    for result in results:
        command = result.command.split("/", 1)[0]
        parser = supported.get(command)
        if parser is None:
            continue
        if command not in grouped:
            grouped[command] = []
            indexes[command] = {}
        label = result.display_name or result.config
        if label not in indexes[command]:
            rows: list[list[float]] = []
            indexes[command][label] = rows
            grouped[command].append((label, rows))
        indexes[command][label].extend(parser(result.response))

    if not grouped:
        raise HudProtocolError("No t24 or t6 results were available for Excel export")

    output = Path(output_path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    fills = ["E2F0D9", "DDEBF7", "FFF2CC", "FCE4D6"]
    header_border = Border(
        left=Side(style="thin", color="7F8C8D"),
        right=Side(style="thin", color="7F8C8D"),
        top=Side(style="thin", color="7F8C8D"),
        bottom=Side(style="thin", color="7F8C8D"),
    )
    data_border = Border(
        left=Side(style="thin", color="D9E1F2"),
        right=Side(style="thin", color="D9E1F2"),
        top=Side(style="thin", color="D9E1F2"),
        bottom=Side(style="thin", color="D9E1F2"),
    )

    configs_per_row = 3
    block_gap = 1
    for command, blocks in grouped.items():
        sheet = workbook.create_sheet(f"{command}测试结果")
        sheet.sheet_view.showGridLines = False
        start_row = 1
        for group_start in range(0, len(blocks), configs_per_row):
            group = blocks[group_start : group_start + configs_per_row]
            max_values = max(len(rows) for _, rows in group)
            for position, (config, rows) in enumerate(group):
                block_index = group_start + position
                column = 1 + position * (1 + block_gap)
                header = sheet.cell(
                    row=start_row,
                    column=column,
                    value=config,
                )
                header.fill = PatternFill("solid", fgColor=fills[block_index % len(fills)])
                header.font = Font(bold=True, color="1F2937")
                header.alignment = Alignment(horizontal="center", vertical="center")
                header.border = header_border
                sheet.column_dimensions[get_column_letter(column)].width = 24
                sheet.row_dimensions[start_row].height = 24

                for value_index, row in enumerate(rows, start=1):
                    cell = sheet.cell(row=start_row + value_index, column=column, value=row[0])
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = data_border
                    cell.number_format = "0.###############"
            start_row += max_values + 2

    image_results = [result for result in results if result.image_path]
    if image_results:
        image_sheet = workbook.create_sheet("图片", 1)
        image_sheet.sheet_view.showGridLines = False
        image_sheet["A1"] = "区域名称"
        image_sheet["B1"] = "图片"
        for cell in image_sheet[1]:
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
            cell.font = Font(bold=True, color="1F2937")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border
        image_sheet.column_dimensions["A"].width = 24
        image_sheet.column_dimensions["B"].width = 42
        image_sheet.freeze_panes = "A2"

        for row_index, result in enumerate(image_results, start=2):
            config_cell = image_sheet.cell(
                row=row_index,
                column=1,
                value=result.display_name or result.config,
            )
            config_cell.alignment = Alignment(horizontal="center", vertical="center")
            config_cell.border = data_border
            image_cell = image_sheet.cell(row=row_index, column=2)
            image_cell.border = data_border
            image = ExcelImage(result.image_path)
            image.width = 240
            image.height = 135
            image_sheet.add_image(image, f"B{row_index}")
            image_sheet.row_dimensions[row_index].height = 105

    workbook.save(output)
    return output


def export_t24_excel(results: list[TestResult], output_path: str) -> Path:
    """Backward-compatible wrapper for callers that export t24 results."""
    return export_results_excel(results, output_path)


class HudClient:
    def __init__(
        self,
        host: str = "127.0.0.1",
        port: int = 5555,
        timeout: float = 60.0,
        response_idle: float = 1.0,
        debug: bool = False,
    ):
        self.host = host
        self.port = port
        self.timeout = timeout
        self.response_idle = response_idle
        self.debug = debug
        self._socket: socket.socket | None = None
        self._buffer = bytearray()
        self._peer_closed = False

    def connect(self) -> None:
        if self._socket is not None and not self._peer_closed:
            return
        if self._socket is not None:
            self._socket.close()
        self._socket = socket.create_connection((self.host, self.port), self.timeout)
        self._socket.settimeout(self.timeout)
        self._peer_closed = False

    def close(self) -> None:
        if self._socket is not None:
            self._socket.close()
            self._socket = None
        self._buffer.clear()
        self._peer_closed = False

    def __enter__(self) -> "HudClient":
        self.connect()
        return self

    def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
        self.close()

    def request(self, command: str) -> str:
        """Send one command and receive through the first '%' terminator."""
        if self._socket is None:
            raise RuntimeError("HUD client is not connected")
        if self._peer_closed:
            self.connect()

        wire_command = command.strip()
        if not wire_command:
            raise ValueError("Command cannot be empty")

        if self.debug:
            print(f"TX {wire_command!r}", flush=True)
        self._socket.sendall(wire_command.encode("utf-8"))
        response = self._receive_message()
        if self.debug:
            print(f"RX {response!r}", flush=True)
        return response

    def _receive_message(self) -> str:
        if self._socket is None:
            raise RuntimeError("HUD client is not connected")

        received_any = bool(self._buffer)
        try:
            while True:
                end = self._buffer.find(TERMINATOR)
                if end >= 0:
                    message = bytes(self._buffer[: end + 1])
                    del self._buffer[: end + 1]
                    return message.decode("utf-8", errors="replace").strip()

                try:
                    chunk = self._socket.recv(4096)
                except socket.timeout:
                    if received_any:
                        message = bytes(self._buffer)
                        self._buffer.clear()
                        return message.decode("utf-8", errors="replace").strip()
                    raise

                if not chunk:
                    if received_any:
                        message = bytes(self._buffer)
                        self._buffer.clear()
                        self._peer_closed = True
                        return message.decode("utf-8", errors="replace").strip()
                    self._peer_closed = True
                    raise ConnectionError("HUD software closed the connection before returning data")

                self._buffer.extend(chunk)
                received_any = True
                self._socket.settimeout(self.response_idle)
        finally:
            if self._socket is not None and not self._peer_closed:
                self._socket.settimeout(self.timeout)

    def switch_config(self, config_name: str) -> None:
        name = config_name.strip()
        if not name:
            raise ValueError("Configuration name cannot be empty")
        if name.lower().endswith(".ini"):
            name = name[:-4]

        response = self.request(f"c-{name}%")
        if response.rstrip("%") != "OK":
            raise HudProtocolError(f"Failed to switch configuration {name!r}: {response}")

    def set_brightness_file(self, file_path: str) -> None:
        response = self.request(f"ssf-{file_path}")
        if response.rstrip("%") != "OK":
            raise HudProtocolError(f"Failed to set brightness file {file_path!r}: {response}")

    def measure(self, command: str) -> str:
        response = self.request(command)
        expected_prefix = command.split("/", 1)[0] + "_Result"
        if response in {"Error0%", "Error1%", "Fail%"}:
            raise HudProtocolError(f"Measurement {command!r} failed: {response}")
        if not response.startswith(expected_prefix):
            raise HudProtocolError(
                f"Unexpected response to {command!r}; expected {expected_prefix!r}, got {response!r}"
            )
        return response


def run_tests(
    client: HudClient,
    configs: list[str],
    commands: list[str],
    switch_delay: float,
) -> list[TestResult]:
    results: list[TestResult] = []
    for config in configs:
        print(f"[{config}] switching configuration...", flush=True)
        client.switch_config(config)
        if switch_delay:
            time.sleep(switch_delay)

        for command in commands:
            print(f"[{config}] running {command}...", flush=True)
            response = client.measure(command)
            results.append(TestResult(config, command, response))
            print(f"[{config}] {response}", flush=True)
    return results


def run_cases(
    client: HudClient,
    cases: list[tuple[str, str]],
    switch_delay: float,
) -> list[TestResult]:
    """Run explicitly paired configuration/command cases."""
    results: list[TestResult] = []
    for config, command in cases:
        print(f"[{config}] switching configuration...", flush=True)
        client.switch_config(config)
        if switch_delay:
            time.sleep(switch_delay)
        print(f"[{config}] running {command}...", flush=True)
        response = client.measure(command)
        results.append(TestResult(config, command, response))
        print(f"[{config}] {response}", flush=True)
    return results


def load_test_plan(path: str) -> list[tuple[str, list[str]]]:
    """Read a literal TEST_PLAN assignment without executing the config file."""
    config_path = Path(path)
    try:
        tree = ast.parse(config_path.read_text(encoding="utf-8"), filename=str(config_path))
    except (OSError, SyntaxError) as exc:
        raise ValueError(f"Cannot read test plan {config_path}: {exc}") from exc

    raw_plan = None
    for node in tree.body:
        if isinstance(node, ast.Assign) and any(
            isinstance(target, ast.Name) and target.id == "TEST_PLAN" for target in node.targets
        ):
            try:
                raw_plan = ast.literal_eval(node.value)
            except (ValueError, TypeError, SyntaxError) as exc:
                raise ValueError("TEST_PLAN must contain literal strings, lists, and tuples only") from exc
            break

    if raw_plan is None:
        raise ValueError(f"TEST_PLAN was not found in {config_path}")
    if not isinstance(raw_plan, (list, tuple)) or not raw_plan:
        raise ValueError("TEST_PLAN must be a non-empty list")

    plan: list[tuple[str, list[str]]] = []
    for index, item in enumerate(raw_plan, start=1):
        if not isinstance(item, (list, tuple)) or len(item) != 2:
            raise ValueError(f"TEST_PLAN item {index} must be (config, [commands])")
        config, commands = item
        if not isinstance(config, str) or not config.strip():
            raise ValueError(f"TEST_PLAN item {index} has an invalid config name")
        if not isinstance(commands, (list, tuple)) or not commands:
            raise ValueError(f"TEST_PLAN item {index} must contain at least one command")
        if not all(isinstance(command, str) and command.strip() for command in commands):
            raise ValueError(f"TEST_PLAN item {index} contains an invalid command")
        plan.append((config.strip(), [command.strip() for command in commands]))
    return plan


def load_plan_options(path: str) -> PlanOptions:
    """Load optional image directories and brightness file from a plan file."""
    config_path = Path(path)
    try:
        tree = ast.parse(config_path.read_text(encoding="utf-8"), filename=str(config_path))
    except (OSError, SyntaxError) as exc:
        raise ValueError(f"Cannot read test plan {config_path}: {exc}") from exc

    values: dict[str, str | None] = {}
    wanted = {"SPARKLE_SOURCE_DIR", "SPARKLE_SAVE_DIR"}
    for node in tree.body:
        if not isinstance(node, ast.Assign):
            continue
        names = [target.id for target in node.targets if isinstance(target, ast.Name)]
        matching = wanted.intersection(names)
        if not matching:
            continue
        try:
            value = ast.literal_eval(node.value)
        except (ValueError, TypeError, SyntaxError) as exc:
            raise ValueError(f"{next(iter(matching))} must be a string or None") from exc
        if value is not None and not isinstance(value, str):
            raise ValueError(f"{next(iter(matching))} must be a string or None")
        for name in matching:
            values[name] = value.strip() if isinstance(value, str) and value.strip() else None

    source = values.get("SPARKLE_SOURCE_DIR")
    destination = values.get("SPARKLE_SAVE_DIR")
    return PlanOptions(source, destination)


def save_sparkle_image(
    config: str,
    source_dir: str,
    save_dir: str,
    sequence: int,
    filename: str | None = None,
) -> Path:
    """Copy Sparkle.jpg to a unique per-measurement destination filename."""
    source = Path(source_dir).expanduser() / "Sparkle.jpg"
    if not source.is_file():
        raise HudProtocolError(f"Sparkle image was not found: {source}")
    destination_dir = Path(save_dir).expanduser()
    destination_dir.mkdir(parents=True, exist_ok=True)
    if filename:
        destination = destination_dir / filename
    else:
        safe_config = re.sub(r"[^0-9A-Za-z._-]+", "_", config).strip("_") or "config"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        destination = destination_dir / f"Sparkle_{safe_config}_{sequence:03d}_{timestamp}.jpg"
    shutil.copy2(source, destination)
    return destination.resolve()


def wait_for_brightness_file(
    file_path: str,
    previous_mtime_ns: int | None,
    timeout: float = 30.0,
    width: int = BRIGHTNESS_WIDTH,
    height: int = BRIGHTNESS_HEIGHT,
) -> Path:
    """Wait until a newly written float32 brightness file reaches its expected size."""
    path = Path(file_path).expanduser()
    expected_size = width * height * FLOAT32_BYTES
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        try:
            stat = path.stat()
        except FileNotFoundError:
            time.sleep(0.2)
            continue
        is_new = previous_mtime_ns is None or stat.st_mtime_ns != previous_mtime_ns
        if is_new and stat.st_size == expected_size:
            return path.resolve()
        time.sleep(0.2)
    actual_size = path.stat().st_size if path.exists() else None
    raise HudProtocolError(
        f"Brightness file was not ready: {path}; expected {expected_size} bytes, "
        f"actual {actual_size if actual_size is not None else 'missing'}"
    )


def run_test_plan(
    client: HudClient,
    plan: list[tuple[str, list[str]]],
    switch_delay: float,
    sparkle_source_dir: str | None = None,
    sparkle_save_dir: str | None = None,
    run_dir: str | None = None,
) -> list[TestResult]:
    results: list[TestResult] = []
    sparkle_sequence = 0
    batch_dir = Path(run_dir).expanduser().resolve() if run_dir else None
    if batch_dir:
        batch_dir.mkdir(parents=True, exist_ok=True)
    for config_index, (config, commands) in enumerate(plan, start=1):
        region_name = f"区域{config_index}"
        print(f"[{config}] switching configuration...", flush=True)
        client.switch_config(config)
        if switch_delay:
            time.sleep(switch_delay)
        for command in commands:
            print(f"[{config}] running {command}...", flush=True)
            if command.startswith("ssf-"):
                configured_path = command[4:].strip()
                if not configured_path:
                    raise HudProtocolError("ssf command is missing the brightness file path")
                brightness_path = (
                    str(batch_dir / f"{region_name}.bin")
                    if batch_dir
                    else configured_path
                )
                brightness_file = Path(brightness_path).expanduser()
                previous_mtime_ns = (
                    brightness_file.stat().st_mtime_ns if brightness_file.exists() else None
                )
                client.set_brightness_file(brightness_path)
                ready_file = wait_for_brightness_file(
                    brightness_path,
                    previous_mtime_ns,
                    timeout=client.timeout,
                )
                print(f"[{config}] brightness bin ready: {ready_file}", flush=True)
                if batch_dir:
                    txt_file = convert_brightness_bin_to_txt(ready_file)
                    print(f"[{config}] brightness txt saved: {txt_file}", flush=True)
                    corrected_file = create_corrected_brightness_txt(
                        ready_file,
                        batch_dir / f"修正{region_name}.txt",
                    )
                    print(
                        f"[{config}] corrected brightness saved: {corrected_file}",
                        flush=True,
                    )
                continue
            response = client.measure(command)
            image_path = None
            image_destination = str(batch_dir) if batch_dir else sparkle_save_dir
            if command.split("/", 1)[0] == "t24" and sparkle_source_dir and image_destination:
                sparkle_sequence += 1
                image_path = str(
                    save_sparkle_image(
                        config,
                        sparkle_source_dir,
                        image_destination,
                        sparkle_sequence,
                        f"{region_name}.jpg" if batch_dir else None,
                    )
                )
                print(f"[{config}] image saved: {image_path}", flush=True)
            results.append(TestResult(config, command, response, image_path, region_name))
            print(f"[{config}] {response}", flush=True)
    return results


def parse_case(value: str) -> tuple[str, str]:
    try:
        config, command = value.rsplit(":", 1)
    except ValueError as exc:
        raise argparse.ArgumentTypeError(
            "Case must use CONFIG:COMMAND format, for example config1:t1"
        ) from exc
    if not config.strip() or not command.strip():
        raise argparse.ArgumentTypeError("Configuration and command cannot be empty")
    return config.strip(), command.strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Switch HUD software configurations and run measurement commands."
    )
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=5555)
    parser.add_argument("--timeout", type=float, default=60.0)
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Print exact transmitted commands and raw received responses.",
    )
    parser.add_argument(
        "--config",
        action="append",
        help="Configuration filename without extension; repeat for multiple configurations.",
    )
    parser.add_argument(
        "--test",
        action="append",
        help="Measurement command such as t1, t3, or t11/10/20; repeat as needed.",
    )
    parser.add_argument(
        "--case",
        action="append",
        type=parse_case,
        help="Paired CONFIG:COMMAND case; repeat as needed, e.g. --case config1:t1.",
    )
    parser.add_argument(
        "--plan",
        help="Python-style test plan file containing a literal TEST_PLAN list.",
    )
    parser.add_argument(
        "--switch-delay",
        type=float,
        default=0.2,
        help="Seconds to wait after switching configuration (default: 0.2).",
    )
    parser.add_argument(
        "--output",
        default="hud_results.xlsx",
        help="Excel output path for t24/t6 results (default: hud_results.xlsx).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.plan and not args.case and not (args.config and args.test):
        print(
            "ERROR: use --plan FILE, --case CONFIG:COMMAND, or both --config and --test",
            file=sys.stderr,
        )
        return 2
    try:
        with HudClient(args.host, args.port, args.timeout, debug=args.debug) as client:
            if args.plan:
                options = load_plan_options(args.plan)
                results = run_test_plan(
                    client,
                    load_test_plan(args.plan),
                    args.switch_delay,
                    options.sparkle_source_dir,
                    options.sparkle_save_dir,
                )
            elif args.case:
                results = run_cases(client, args.case, args.switch_delay)
            else:
                results = run_tests(client, args.config, args.test, args.switch_delay)
        if any(result.command.split("/", 1)[0] in {"t24", "t6"} for result in results):
            output = export_results_excel(results, args.output)
            print(f"Excel saved: {output}", flush=True)
    except (OSError, HudProtocolError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
